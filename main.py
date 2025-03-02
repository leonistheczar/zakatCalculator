import time
import json
import os
import openpyxl 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# NISAB AS FOR 25th FEB, 2025
GOLD_NISAB = 2343764  # PKR
SILVER_NISAB = 178689  # PKR

def get_user_input():
    print("Check your eligibility for Zakat: ")
    userName = input("Enter your name: ")
    cash = int(input("Enter the value of cash in hand: "))
    bank_savings = int(input("Enter the value of cash in bank: "))
    gold = int(input("Enter the value of gold in possession: "))
    silver = int(input("Enter the value of silver in possession: "))
    company_assets = int(input("Enter the value of stocks or shares in possession: "))
    agri_products = int(input("Enter the value of agricultural products in possession: "))
    land = int(input("Enter the value of land in possession for sale: "))
    debts = int(input("Enter the value of any debts you owe: "))
    return cash, gold, silver, company_assets, agri_products, land, bank_savings, debts, userName

def calculate_total_assets(cash, gold, silver, company_assets, agri_products, land, bank_savings, debts):
    return cash + gold + silver + company_assets + agri_products + land + bank_savings - debts

def check_eligibility(total_assets):
    return total_assets >= SILVER_NISAB

def calculate_zakat(total_assets):
    return total_assets * 0.025

def export_to_txt(data):
    with open("zakat_report.txt", "a") as file:  # Append mode
        file.write("\n----- Zakat Calculation Report -----\n")
        for key, value in data.items():
            file.write(f"{key}: {value:,} PKR\n" if isinstance(value, (int, float)) else f"{key}: {value}\n")
        file.write("\n")

def export_to_json(data):
    file_exists = os.path.exists("zakat_report.json")
    
    if file_exists:
        with open("zakat_report.json", "r") as file:
            try:
                existing_data = json.load(file)
            except json.JSONDecodeError:
                existing_data = []
    else:
        existing_data = []

    existing_data.append(data)

    with open("zakat_report.json", "w") as file:
        json.dump(existing_data, file, indent=4)
def export_to_excel(data):
    file_name = "zakat_report.xlsx"

    # Check if the file exists
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)  # Load existing workbook
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        # Create headers if file is new
        sheet.append([
            "Name", "Cash", "Bank Savings", "Gold", "Silver", "Company Assets", 
            "Agricultural Products", "Land", "Debts", "Total Assets", "Eligibility", "Zakat Amount"
        ])

    # Append new data
    sheet.append([
        data["Name"], data["Cash"], data["Bank Savings"], data["Gold"], data["Silver"],
        data["Company Assets"], data["Agricultural Products"], data["Land"],
        data["Debts"], data["Total Assets"], data["Eligibility"], data["Zakat Amount"]
    ])

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Save the Excel file
    wb.save(file_name)
    print("\n✅ Data has been successfully saved to Excel!")

def main():
    print("\n----- Zakat Calculator Project -----\n")
    cash, gold, silver, company_assets, agri_products, land, bank_savings, debts, userName = get_user_input()
    total_assets = calculate_total_assets(cash, gold, silver, company_assets, agri_products, land, bank_savings, debts)
    
    print("\nCalculating total assets...")
    time.sleep(2)
    print(f"Total assets: {total_assets:,} PKR")

    eligibility = "Eligible" if check_eligibility(total_assets) else "Not Eligible"
    zakat = calculate_zakat(total_assets) if eligibility == "Eligible" else 0
    
    if eligibility == "Eligible":
        print("\nPlease wait...")
        time.sleep(2)
        print(f"{userName} is eligible to pay Zakat.")
        print("\nCalculating Zakat now...")
        time.sleep(2)
        print(f"Zakat amount to be paid: {zakat:,} PKR")
    else:
        print("You are not eligible to pay Zakat as your wealth is lower than the Nisab threshold.")

    # Prepare data for export
    zakat_data = {
        "Name": userName,
        "Cash": cash,
        "Bank Savings": bank_savings,
        "Gold": gold,
        "Silver": silver,
        "Company Assets": company_assets,
        "Agricultural Products": agri_products,
        "Land": land,
        "Debts": debts,
        "Total Assets": total_assets,
        "Eligibility": eligibility,
        "Zakat Amount": zakat
    }

    # Export to all formats with append functionality
    export_to_txt(zakat_data)
    export_to_excel(zakat_data)
    export_to_json(zakat_data)  # JSON export added
    
    print("\nData has been saved successfully and appended to existing files!")

if __name__ == "__main__":
    main()
